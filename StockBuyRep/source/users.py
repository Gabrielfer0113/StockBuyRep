import json
from pathlib import Path
from datetime import datetime
from time import sleep
import math
import shutil
import sqlite3

from log import LogFileMixin

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import PySimpleGUI as sg
import pandas as pd


# PATHS
KONEMETAL_DATABASE = str(Path(__file__).parent.parent / "configuration\\database\\konemetal.sqlite3")
KONEMETAL_BACKUP = Path(__file__).parent.parent / "configuration\\backup" / f'konemetal_{str(datetime.now().date())[2:]}.sqlite3'
DAVIK_DATABASE = str(Path(__file__).parent.parent / "configuration\\database\\davik.sqlite3")
DAVIK_BACKUP = Path(__file__).parent.parent / "configuration\\backup" / f'davik_{str(datetime.now().date())[2:]}.sqlite3'
PREFERENCES = Path(__file__).parent.parent / "configuration\\preferences.json"
SUPPLIERS = Path(__file__).parent.parent / "configuration\\suppliers.json"
MATERIAL_STANDARDS = Path(__file__).parent.parent / "configuration\\material_standards.json"
COLOR_MAP = Path(__file__).parent.parent / "configuration\\color_map.json"

class User(LogFileMixin):
    def __init__(self, name:str, password:str, area:str, user_pc:str, 
    config:bool, themes:bool, report_tools:bool, receive_mp:bool, delete_mp:bool, 
    register_mp:bool, edit_mp:bool, mp_filter:bool, debug_tools:bool, 
    email_tools:bool, historic:bool, theme_name:str, button_color:str, 
    pressed_button_color:str, second_color:str, text_color:str, font:str, 
    font_size:int, gauge_list:list(), market_gauge_list:list(), material_list:list(), 
    market_material_list:list(), standards_list:list(), market_standards_list:list(),
    forging_weight_list:dict) -> None:
        self.name = name.capitalize()
        self.password = password
        self.area = area
        self.PC = user_pc

        self.config_acess = config
        self.themes_acess = themes
        self.report_tools_acess = report_tools
        self.receive_mp_acess = receive_mp
        self.delete_mp_acess = delete_mp
        self.register_mp_acess = register_mp
        self.edit_mp_acess = edit_mp
        self.mp_filter_acess = mp_filter
        self.debug_tools_acess = debug_tools
        self.email_tools_acess = email_tools
        self.historic_acess = historic

        self.theme_name = theme_name
        self.button_color = button_color
        self.pressed_button_color = pressed_button_color
        self.second_color = second_color
        self.text_color = text_color
        self.font = font
        self.font_size = font_size

        self.gauge_list = gauge_list
        self.market_gauge_list = market_gauge_list
        self.material_list = material_list
        self.market_material_list = market_material_list
        self.standards_list = standards_list
        self.market_standards_list = market_standards_list
        self.forging_weight_list = forging_weight_list
        
        self.debug_status = False
        self.last_filter = 'id'
        self.last_filter_order = 'ASC'
        self.last_report = list()

    def changePreferences(self, theme:str, button_color:str, 
    pressed_button_color:str, second_color:str, text_color:str, font:str, 
    font_size:str) -> None: 

        with open(PREFERENCES, 'r') as file:
            report_file = json.load(file)
            report_file[self.name]['Theme'] = theme
            report_file[self.name]['button_color'] = button_color
            report_file[self.name]['pressed_button_color'] = pressed_button_color
            report_file[self.name]['second_color'] = second_color
            report_file[self.name]['text_color'] = text_color
            report_file[self.name]['font'] = font
            report_file[self.name]['font_size'] = font_size
        with open(PREFERENCES, 'w') as file:
            json.dump(report_file, file, indent=4)

    def calculateWeight(self, formato, d1=0, d2=0, material='material', metros=0, interna=False):
        if material == 'aco':
            coeficiente_red = 0.0062
            coeficiente_quad = 0.0079
            coeficiente_sext = 0.0068
        elif material == 'latao':
            coeficiente_red = 0.006683
            coeficiente_quad = 0 # nao vendemos
            coeficiente_sext = 0.007359      
        elif material == 'aluminio':
            coeficiente_red = 0.00212058
            coeficiente_quad = 0.002728
            coeficiente_sext = 0.002344    

        if formato == 'redondo':
            valor = (d1**2)*coeficiente_red
        elif formato == 'quadrado':
            valor = (d1*d2)*coeficiente_quad
        elif formato == 'sextavado':
            valor = (d1**2)*coeficiente_sext         
        elif formato == 'tubo':
            valor = ((d1**2)*coeficiente_red) - ((d2**2)*coeficiente_red)

        if interna:
            valor = valor * metros
            return math.ceil(valor)
        else:
            valor = '{:.3f}'.format(valor)
            valor = valor.replace('.', ',')
            return f'{valor}kg/M'       

    def polegada_mm(self, string:str, interna=False) -> int:  # Converte polegada em mm.
        """
        :param string: Passa a fração em formato string por exemplo ("1-1/2")
        :return: retorna valor float em mm
        """
        fracao = string.strip('"')
        if '-' in string:
            fracao = fracao.split('-')
            inteiro = int(fracao[0])
            fracao = fracao[1].split('/')
            dividendo = int(fracao[0])
            divisor = int(fracao[1])
            valor_mm = (25.4 * (inteiro + (dividendo / divisor)))
        elif '/' in string:
            fracao = fracao.split('/')
            dividendo = int(fracao[0])
            divisor = int(fracao[1])
            valor_mm = (25.4 * (dividendo / divisor))
        else:
            valor_mm = int(fracao) * 25.4
        if interna:
            return valor_mm
        valor_mm = str('{:.2f}mm'.format(valor_mm))
        return valor_mm

    def mm_polegada(self, string:str) -> str:
        valor = string.replace(',', '.')
        valor_mm = float(valor)
        pol_inteira = valor_mm // 25.4
        if valor_mm % 25.4 == 0:
            polegada = f'{int(pol_inteira)}"'
            return polegada
        else:
            valor_mm = valor_mm - (pol_inteira * 25.4)
            dividendo = int((valor_mm / 25.4) * 128)
            divisor = 128

            while dividendo % 2 == 0 and dividendo != 0:
                dividendo /= 2
                divisor /= 2

            if pol_inteira:
                polegada = f'{int(pol_inteira)}-{int(dividendo)}/{int(divisor)}"'
                return polegada
            else:
                polegada = f'{int(dividendo)}/{int(divisor)}"'
                return polegada

    def forgedWeight(self, valor:int) -> dict:
        return {
            'JL2':{
                'Quilo':'Ø3/4" x {:.2f}Kg'.format( 0.129*int(valor)),
                'Unidade':f'{int(valor/0.129)} un'
            },
            'JL3':{
                'Quilo':'Ø1" x {:.2f}Kg'.format( 0.255 * int(valor)),
                'Unidade':f'{int(valor/0.255)} un'
            },
            'JL4':{
                'Quilo':'Ø1-1/8" x {:.2f}Kg'.format( 0.407 * int(valor)),
                'Unidade':f'{int(valor/0.407)} un'
            },
            'JL5':{
                'Quilo':'Ø1-1/4" x {:.2f}Kg'.format( 0.61 * int(valor)),
                'Unidade':f'{int(valor/0.61)} un'
            },
            'JL6-Flange':{
                'Quilo':'Ø1-1/2" x {:.2f}Kg'.format( 1.18 * int(valor)),
                'Unidade':f'{int(valor/1.18)} un'
            },
            'TL2':{
                'Quilo':'Ø1" x {:.2f}Kg'.format( 0.233 * int(valor)),
                'Unidade':f'{int(valor/0.233)} un'
            },
            'TL3':{
                'Quilo':'Ø1-1/4" x {:.2f}Kg'.format( 0.452 * int(valor)),
                'Unidade':f'{int(valor/0.452)} un'
            },
            'TL5':{
                'Quilo':'Ø1-3/4" x {:.2f}Kg'.format( 0.952 * int(valor)),
                'Unidade':f'{int(valor/0.952)} un'
            },
            'TLCH36':{
                'Quilo':'Ø1-3/4" x {:.2f}Kg'.format( 1.45 * int(valor)),
                'Unidade':f'{int(valor/0.145)} un'
            },
            'CZ4':{
                'Quilo':'Ø1-1/2" x {:.2f}Kg'.format( 0.615 * int(valor)),
                'Unidade':f'{int(valor/0.615)} un'
            },
            'CZ6':{
                'Quilo':'Ø2" x {:.2f}Kg'.format( 2.08 * int(valor)),
                'Unidade':f'{int(valor/2.08)} un'
            }
        }

    def convert_mm_kg(self, milimetros, bitola, norma, dimensao):
        if bitola == "R":
            formato = 'redondo'
        elif bitola == "S":
            formato = 'sextavado'
        elif bitola == "Q" or bitola == 'BC':
            formato = 'quadrado'
        elif bitola == "TN" or bitola == "TC" or bitola == "TB":
            formato = 'tubo'
        else:
            formato = 'x'

        if norma in 'ABCDGWZRIACFLMECTAM':
            material = 'aco'
        elif 'Q' in norma:
            material = 'aluminio'
        elif 'J' in norma:
            material = 'latao'

        if '"' in dimensao or '/' in dimensao:
            dimensao_convertida_mm = self.polegada_mm((dimensao), interna=True)
            dimensao1 = dimensao2 = dimensao_convertida_mm
        elif 'X' in dimensao:
            dimensao = dimensao.split('X')
            if '"' in dimensao[0] or '/' in dimensao[0]:
                dimensao[0] = self.polegada_mm((dimensao[0]), interna=True)
            if '"' in dimensao[1] or '/' in dimensao[1]:
                dimensao[1] = self.polegada_mm((dimensao[1]), interna=True)
            dimensao[0] = dimensao[0].replace(',','.')
            dimensao[1] = dimensao[1].replace(',','.')
            dimensao1 = float(dimensao[0])
            dimensao2 = float(dimensao[1])                
        else:
            if not formato == 'x':
                dimensao = float(dimensao.replace(',','.'))
                dimensao1 = dimensao2 = dimensao
        metros = (int(milimetros)/1000)
        if formato == 'redondo' or formato == 'quadrado' or formato == 'sextavado' or formato == 'tubo':
            kg = self.calculateWeight(formato, d1=dimensao1, d2=dimensao2, material=material, metros=metros, interna=True)
            valor_kgs = kg
        else:
            valor_kgs = milimetros
        return valor_kgs
    
    def update_portifolio(self):
        ...

    def updateReport(self) -> list:
        if self.area == 'Konemetal':
            conn = sqlite3.connect(KONEMETAL_DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
            SELECT Bitola, Material, Norma, Dimensao, Fabricacao, Kg_Solicitado,
            Documento, Data_Entrega_Peca, Urgente, Comprado, id

            FROM Konemetal_Main
            ''')
            material_list = cursor.fetchall()
            conn.close()

            for i in range(len(material_list)):
                material_list[i] = list(material_list[i])
                # material_list[i][-4] = (material_list[i][-4])
                
        elif self.area == 'DaviK':
            conn = sqlite3.connect(str(DAVIK_DATABASE))
            cursor = conn.cursor()
            cursor.execute('''
            SELECT Data_Solicitado, Bitola, Material, Norma, Dimensao, Fabricacao, 
            Kg_Solicitado, Urgente, Comprado, id

            FROM DaviK_Main
            ''')
            material_list = cursor.fetchall()
            conn.close()

            for i in range(len(material_list)):
                material_list[i] = list(material_list[i])
                # material_list[i][0] = self.dataFormater(material_list[i][0])        

        elif self.area == 'Compras':
            conn = sqlite3.connect(KONEMETAL_DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT 
                    Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Concatenado,
                    Kg_Solicitado, COALESCE(Data_Entrega_Peca, "N.Comprado") AS Data_Entrega_Peca, COALESCE(Valor, 0.0) AS Valor, Data_Previsao,
                    Urgente, Comprado, id

                FROM Konemetal_Main
            ''')
            material_list_kone = cursor.fetchall()
            conn.close()

            for i in range(len(material_list_kone)):
                material_list_kone[i] = list(material_list_kone[i])
                # try:
                #     material_list_kone[i][-4] = self.dataFormater(material_list_kone[i][-4])
                # except:
                #     pass
                # material_list_kone[i][-6] = self.dataFormater(material_list_kone[i][-6])
                material_list_kone[i].insert(0,'Konemetal')


            conn = sqlite3.connect(str(DAVIK_DATABASE))
            cursor = conn.cursor()
            cursor.execute('''
            SELECT
                Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Concatenado,
                Kg_Solicitado, COALESCE(Data_Solicitado, "N.Comprado") AS Data_Solicitado, COALESCE(Valor, 0.0) AS Valor, Data_Previsao, 
                Urgente, Comprado, id

            FROM DaviK_Main
            ''')
            material_list_davik = cursor.fetchall()
            conn.close()

            for i in range(len(material_list_davik)):
                material_list_davik[i] = list(material_list_davik[i])
                # try:
                #     material_list_davik[i][-4] = self.dataFormater(material_list_davik[i][-4])
                # except:
                #     pass
                # material_list_davik[i][-6] = self.dataFormater(material_list_davik[i][-6])
                material_list_davik[i].insert(0,'DaviK')


            material_list = material_list_davik + material_list_kone
            # print(material_list_davik)
        else:
            raise KeyError('Area nao encontrada (self.area)')
        return material_list

    def catchValues(self, line:int, historic:bool=False, requester:str=None) -> dict:
        if self.area == 'Konemetal' or requester == 'Konemetal':
            conn = sqlite3.connect(KONEMETAL_DATABASE)
            cursor = conn.cursor()
            if historic:
                cursor.execute(f'''
                SELECT 

                id,
                Bitola, Material, Norma, Dimensao, Fabricacao, 
                Fornecedor, Valor,
                Kg_Solicitado, Kg_Comprado, Kg_Recebido,
                Data_Entrega_Peca, Data_Solicitado,  Data_Comprado,  Data_Previsao, Data_Recebido,
                Ordem_Compra, Nota_Fiscal,
                Documento, Codigo_Peca,
                Solicitante

                FROM 
                    Konemetal_Historic
                WHERE 
                    id = {line}
                ''')
                values = cursor.fetchall()
                conn.close()

                material_info= {
                    'id':values[0][0],
                    
                    'Bitola':values[0][1],
                    'Material':values[0][2], 
                    'Norma':values[0][3], 
                    'Dimensao':values[0][4],  
                    'Fabricacao':values[0][5],

                    'Fornecedor':values[0][6], 
                    'Valor':values[0][7],
                    
                    'Kg_Solicitado':values[0][8], 
                    'Kg_Comprado':values[0][9],
                    'Kg_Recebido':values[0][10],
                    
                    'Data_Entrega_Peca':self.dataFormater(values[0][11]), 
                    'Data_Solicitado':self.dataFormater(values[0][12]),  
                    'Data_Comprado':self.dataFormater(values[0][13]),  
                    'Data_Previsao':self.dataFormater(values[0][14]),
                    'Data_Recebido':self.dataFormater(values[0][15]),

                    'Ordem_Compra':values[0][16],
                    'Nota_Fiscal':values[0][17],

                    'Documento':values[0][18], 
                    'Codigo_Peca':values[0][19],
                    'Solicitado_por':values[0][20],

                    "Comprado":False,
                    "Urgente":False,
                    "Milimetros":False
                }
            else:
                cursor.execute(f'''
                SELECT 

                id,
                Bitola, Material, Norma, Dimensao,  Milimetros, Fabricacao, 
                Fornecedor, Valor,
                Kg_Solicitado, Kg_Comprado,

                strftime('%d/%m/%Y', Data_Entrega_Peca) as Data_Entrega_Peca,
                strftime('%d/%m/%Y', Data_Solicitado) as Data_Solicitado,
                strftime('%d/%m/%Y', Data_Comprado) as Data_Comprado,
                strftime('%d/%m/%Y', Data_Previsao) as Data_Previsao,
 
                Ordem_Compra,
                Urgente, Comprado,
                Documento, Codigo_Peca,
                Solicitante

                FROM Konemetal_Main
                WHERE id = {line}
                ''')
                
                values = cursor.fetchall()
                conn.close()

                material_info= {
                    'id':values[0][0],
                    'Bitola':values[0][1],
                    'Material':values[0][2], 
                    'Norma':values[0][3], 
                    'Dimensao':values[0][4],  
                    'Milimetros':values[0][5], 
                    'Fabricacao':values[0][6], 
                    'Fornecedor':values[0][7], 
                    'Valor':values[0][8],
                    'Kg_Solicitado':values[0][9], 
                    'Kg_Comprado':values[0][10],
                    'Data_Entrega_Peca':values[0][11], 
                    'Data_Solicitado':values[0][12],  
                    'Data_Comprado':values[0][13],  
                    'Data_Previsao':values[0][14],
                    'Ordem_Compra':values[0][15],
                    'Urgente':values[0][16], 
                    'Comprado':values[0][17],
                    'Documento':values[0][18], 
                    'Codigo_Peca':values[0][19],
                    'Solicitado_por':values[0][20],

                    'Kg_Recebido':'',
                    'Data_Recebido':'',
                    'Nota_Fiscal':''
                }
        elif self.area == 'DaviK' or requester == 'DaviK':
            conn = sqlite3.connect(DAVIK_DATABASE)
            cursor = conn.cursor()
            if historic:
                cursor.execute(f'''
                SELECT 
                    id,
                    Bitola, Material, Norma, Dimensao, Fabricacao, 
                    Fornecedor, Valor, 
                    Kg_Solicitado, Kg_Comprado, Kg_Recebido,

                    strftime('%d/%m/%Y', Data_Solicitado) AS Data_Recebido,
                    strftime('%d/%m/%Y', Data_Comprado) AS Data_Comprado,
                    strftime('%d/%m/%Y', Data_Previsao) AS Data_Previsao,
                    strftime('%d/%m/%Y', Data_Recebido) AS Data_Recebido,

                    Ordem_Compra, Nota_Fiscal,
                    Solicitante

                FROM Davik_Historic
                WHERE id = {line}
                ''')

                values = cursor.fetchall()
                conn.close()

                material_info = {
                    "id":values[0][0],

                    "Bitola":values[0][1],
                    "Material":values[0][2],
                    "Norma":values[0][3],
                    "Dimensao":values[0][4],
                    "Fabricacao":values[0][5],

                    "Fornecedor":values[0][6],
                    "Valor":values[0][7],

                    "Kg_Solicitado":values[0][8],
                    "Kg_Comprado":values[0][9],
                    'Kg_Recebido':values[0][10],

                    "Data_Solicitado":values[0][11],
                    "Data_Comprado":values[0][12],
                    "Data_Previsao":values[0][13],
                    "Data_Recebido":values[0][14],

                    "Ordem_Compra":values[0][15],
                    "Nota_Fiscal":values[0][16],
                    'Solicitado_por':values[0][17],

                    "Comprado":False,
                    "Urgente":False,
                    "Milimetros":False
                }
            else:
                cursor.execute(f'''
                SELECT 
                    id,
                    Bitola, Material, Norma, Dimensao, Milimetros, Fabricacao, 
                    Fornecedor, Valor, 
                    Kg_Solicitado, Kg_Comprado,

                    strftime('%d/%m/%Y', Data_Solicitado) as Data_Solicitado,
                    strftime('%d/%m/%Y', Data_Comprado) as Data_Comprado,
                    strftime('%d/%m/%Y', Data_Previsao) as Data_Previsao,

                    Ordem_Compra, Urgente, Comprado,
                    Solicitante

                FROM DaviK_Main
                WHERE id = {line}
                ''')

                values = cursor.fetchall()
                conn.close()

                material_info = {
                    "id":values[0][0],

                    "Bitola":values[0][1],
                    "Material":values[0][2],
                    "Norma":values[0][3],
                    "Dimensao":values[0][4],
                    "Milimetros":values[0][5],
                    "Fabricacao":values[0][6],

                    "Fornecedor":values[0][7],
                    "Valor":values[0][8],

                    "Kg_Solicitado":values[0][9],
                    "Kg_Comprado":values[0][10],

                    "Data_Solicitado":values[0][11],
                    "Data_Comprado":values[0][12],
                    "Data_Previsao":values[0][13],

                    "Ordem_Compra":values[0][14],
                    "Urgente":values[0][15],
                    "Comprado":values[0][16],
                    'Solicitado_por':values[0][17],

                    "Kg_Recebido":'',
                    "Data_Recebido":'',
                    "Nota_Fiscal":'',
                }
        return material_info

    def coloringRows(self, data:list, user_text_color:str, user_background_color:str) -> list:
        colored_rows = list()
        counter = 0
        for item in data:
            # print('\n',item)
            if item[-3]:  # Urgente
                text_color = 'red'
            else:
                text_color = user_text_color

            if item[-2]:  # Comprado
                background_color = 'green'
                if self.area == 'Konemetal':    
                    data_entrega = item[-4]
                elif self.area == 'DaviK':    
                    data_entrega = item[0]
                elif self.area == 'Compras':    
                    data_entrega = item[-4]
                
                data_entrega = datetime.strptime(data_entrega, '%Y-%m-%d')
                TODAY = datetime.strptime(datetime.now().date().strftime('%d/%m/%Y'), '%d/%m/%Y')

                if data_entrega < TODAY:
                    background_color = 'orange'
                    text_color = 'red'
            else:
                background_color = user_background_color

            colored_rows.append(
                (counter, text_color, background_color)
            )
            counter += 1
        return colored_rows

    def marketStandard(self, materiais:dict) -> dict:
        with open(MATERIAL_STANDARDS, 'r', encoding='utf-8') as file:
            standard = json.load(file)
        
        materiais['Bitola'] = standard['Bitola'][materiais['Bitola']]
        materiais['Material'] = standard['Material'][materiais['Material']]
        materiais['Norma'] = standard['Norma'][materiais['Norma']]

        return materiais

    def dataFormater(self, date:str) -> str:
        try:
            date = datetime.strptime(date, '%Y-%m-%d')
            date = date.strftime('%d/%m/%Y')
            return date
        except:
            return None

    def dashbordData(self):
        graph_data = dict()
        top5_names = list()
        top5_values = list()
        norma_names = list()
        norma_values = list()
        bitola_names = list()
        bitola_values = list()
        fabricacao_names = list()
        fabricacao_values = list()

        conn = sqlite3.connect(DAVIK_DATABASE)
        cursor = conn.cursor()

        # Top 5
        cursor.execute(f'''
            SELECT Bitola, Material, Norma, Dimensao, Fabricacao, SUM(Kg_Solicitado) as total_quilos
            FROM Davik_Main
            GROUP BY Bitola, Material, Norma, Dimensao
            ORDER BY total_quilos DESC''')
        itens_list = cursor.fetchall()
        for item in itens_list:
            bitola, material, norma, dimensao, fabricacao, kg = item
            top5_names.append(f'{bitola}{material}{norma} {dimensao} {fabricacao[0].capitalize()}')
            top5_values.append(kg)
        graph_data['top5_names'] = top5_names[0:5]
        graph_data['top5_values'] = top5_values[0:5]

        # Norma
        cursor.execute(f'''
            SELECT
                Norma,
                (SUM(Kg_Solicitado) * 100.0 / (SELECT SUM(Kg_Solicitado) FROM Davik_Main)) as percentual_quilos
            FROM Davik_Main
            GROUP BY Norma
            ORDER BY percentual_quilos DESC''')
        itens_list = cursor.fetchall()
        for item in itens_list:
            norma, kg_percent = item
            norma_names.append(norma)
            norma_values.append(kg_percent)
        graph_data['norma_names'] = norma_names
        graph_data['norma_values'] = norma_values

        # Bitola
        cursor.execute(f'''
            SELECT
                Bitola,
                (SUM(Kg_Solicitado) * 100.0 / (SELECT SUM(Kg_Solicitado) FROM Davik_Main)) as percentual_quilos
            FROM Davik_Main
            GROUP BY Bitola
            ORDER BY percentual_quilos DESC''')
        itens_list = cursor.fetchall()
        for item in itens_list:
            Bitola, kg_percent = item
            bitola_names.append(Bitola)
            bitola_values.append(kg_percent)
        graph_data['bitola_names'] = bitola_names
        graph_data['bitola_values'] = bitola_values

        # Fabricacao
        cursor.execute(f'''
            SELECT
                Fabricacao,
                (SUM(Kg_Solicitado) * 100.0 / (SELECT SUM(Kg_Solicitado) FROM Davik_Main)) as percentual_quilos
            FROM Davik_Main
            GROUP BY Fabricacao
            ORDER BY percentual_quilos DESC''')
        itens_list = cursor.fetchall()
        for item in itens_list:
            fabricacao, kg_percent = item
            fabricacao_names.append(fabricacao)
            fabricacao_values.append(kg_percent)
        graph_data['fabricacao_names'] = fabricacao_names
        graph_data['fabricacao_values'] = fabricacao_values

        # Kg Total
        cursor.execute(f"""
        SELECT SUM(Kg_Solicitado) as soma_total_quilos
        FROM Davik_Main""")
        kg_total = cursor.fetchall()
        graph_data['kg_total'] = kg_total
        conn.close()

        return graph_data

    def popupColorChooser(self, msg_used:str):
        """
        msg_used: Where the color will be used
        :return: A dict {'Cor':color_name, 'Codigo':color_code}
        """
        with open(COLOR_MAP,'r') as file:
            color_map = json.load(file)

        button_size = (1, 1)

        # button_size = (None,None)         # for very compact buttons

        def ColorButton(color):
            """
            A User Defined Element - returns a Button that configured in a certain way.
            :param color: Tuple[str, str] ( color name, hex string)
            :return: sg.Button object
            """
            return sg.B(button_color=('white', color[1]), pad=(0, 0), size=button_size, key=color, tooltip=f'{color[0]}:{color[1]}', border_width=0)

        num_colors = len(list(color_map.keys()))
        row_len = 40

        grid = [[ColorButton(list(color_map.items())[c + j * row_len]) for c in range(0, row_len)] for j in range(0, num_colors // row_len)]
        grid += [[ColorButton(list(color_map.items())[c + num_colors - num_colors % row_len]) for c in range(0, num_colors % row_len)]]

        layout = [[sg.Text(msg_used, font='Def 18')]] + grid + \
                [[sg.Button('OK'), sg.T(size=(30, 1), key='-OUT-')]]

        window = sg.Window('Window Title', layout, no_titlebar=True, grab_anywhere=True, keep_on_top=True, use_ttk_buttons=True)
        color_chosen = None
        while True:  # Event Loop
            event, values = window.read()
            if event in (sg.WIN_CLOSED, 'OK'):
                if event == sg.WIN_CLOSED:
                    color_chosen = None
                break
            window['-OUT-'](f'Cor: {event[0]} codigo: {event[1]}')
            color_chosen = {'Cor':event[0], 'Codigo':event[1]}
        window.close()

        return color_chosen

    def filterReport(self, filter_by:str, filter_key:str, historic:bool=False):

        print('filter_by por: ',filter_by)
        print('filter_key: ',filter_key)
        print('historic: ',historic)

        if self.area == 'DaviK':
            conn = sqlite3.connect(DAVIK_DATABASE)
            cursor = conn.cursor()
            if historic:
                sql_query = f'''
                    SELECT 
                        strftime('%d/%m/%Y', Data_Recebido) AS data_formatada,
                        Fornecedor, 
                        Ordem_Compra, 
                        Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                        Kg_Recebido, 
                        Valor, 
                        id
                    FROM 
                        Davik_Historic
                    WHERE
                        {filter_by} = '{filter_key}'
                '''
            else:
                sql_query = f'''
                    SELECT 
                        Data_Solicitado, Bitola, Material, Norma, Dimensao, Fabricacao, 
                        Kg_Solicitado, Urgente, Comprado, id
                    FROM 
                        DaviK_Main
                    WHERE
                        {filter_by} = '{filter_key}'
                '''
            cursor.execute(sql_query)
            item_list = cursor.fetchall()
            conn.close()
        elif self.area == 'Konemetal':
            conn = sqlite3.connect(KONEMETAL_DATABASE)
            cursor = conn.cursor()
            if historic:
                sql_query = f'''
                    SELECT 
                        strftime('%d/%m/%Y', Data_Recebido) AS data_formatada,
                        Fornecedor, 
                        Ordem_Compra, 
                        Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                        Kg_Recebido, 
                        Valor, 
                        id
                    FROM 
                        Konemetal_Historic
                    WHERE
                        {filter_by} = '{filter_key}'
                '''
            else:
                sql_query = f'''
                    SELECT 
                        Bitola, Material, Norma, Dimensao, Fabricacao, Kg_Solicitado,
                        Documento, Data_Entrega_Peca, Urgente, Comprado, id
                    FROM 
                        Konemetal_Main
                    WHERE
                        {filter_by} = '{filter_key}'
                '''

            print(sql_query)
            cursor.execute(sql_query)
            item_list = cursor.fetchall()
            conn.close()
        return item_list

    def configurationSetter(self, category:str, internal_std:str, market_std:str):
        with open(str(MATERIAL_STANDARDS), 'r', encoding='UTF-8') as file:
            report = json.load(file)

        report[category][internal_std] = market_std

        with open(str(MATERIAL_STANDARDS), 'w', encoding='UTF-8') as file:
            json.dump(report, file, indent=4)

    def suppliers_parameter_config(self, material_config:bool, empresa:str, 
    entrega:bool=None, kg_min:int=None, val_min:int=None, item_min:str=None, 
    tax:str=None, seller:str=None, email:str=None, material_area:str=None, 
    material_parameter:str=None):
        with open(SUPPLIERS, 'r', encoding='UTF-8') as file:
            archive = json.load(file)

        if material_config:
            archive[empresa]['Normas'][material_area].append(material_parameter)
        else:
            archive[empresa]['Frete']['Entrega'] = entrega
            archive[empresa]['Frete']['Kg_minimo'] = kg_min
            archive[empresa]['Frete']['Valor_minimo'] = val_min
            archive[empresa]['Frete']['Kg_minimo_por_item'] = item_min
            archive[empresa]['Frete']['Taxa'] = tax
            archive[empresa]['Nome_vendedor'] = seller
            archive[empresa]['email'] = email

        with open(SUPPLIERS, 'w', encoding='UTF-8') as file:
            json.dump(archive, file, indent=4)

    def sortedRowsTables(self, report:list, column:str, order:str) -> list:

        if order == 'ASC':
            order = True
        else:
            order = False

        material_list = sorted(report, key=lambda x: x[column], reverse=order)

        return material_list

class DavikUser(User, LogFileMixin):
    def registerReport(self, bitola:str, material:str, norma:str, dimensao:str, 
        milimetros:int, fabricacao:str, user_name) -> None:

        valor_kg = self.convert_mm_kg(float(milimetros), bitola.upper(), norma.upper(), dimensao.upper())

        conn = sqlite3.connect(DAVIK_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO 
                DaviK_Main 
                (
                    Bitola, Material, Norma, Dimensao, Milimetros, Fabricacao, 
                    Fornecedor, Valor,
                    Kg_Solicitado, Kg_Comprado,
                    Data_Solicitado, Data_Comprado, Data_Previsao,
                    Ordem_Compra,
                    Urgente, Comprado, Solicitante
                ) 
            VALUES 
                (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
            """,
            (
                bitola.upper(), material.upper(), norma.upper(), dimensao.upper(), int(milimetros), fabricacao.upper(), 
                None, None,
                valor_kg, None, 
                datetime.now().date(), None, None, 
                None,
                0, 0,
                user_name
            )
        )
        conn.commit()
        conn.close()

    def saveFile(self) -> None:
        shutil.copy(
            DAVIK_DATABASE, DAVIK_BACKUP
        )
        
    def editReport(self, index:int, bitola:str, material:str, norma:str, 
    dimensao:str, milimetros:int, fabricacao:str, comprado, urgente:int) -> None:
       
        quilos = self.convert_mm_kg(milimetros, bitola, norma, dimensao)

        conn = sqlite3.connect(DAVIK_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            f"""
            UPDATE DaviK_Main 
            SET 
                Bitola = ?,
                Material = ?,
                Norma = ?,
                Dimensao = ?,
                Milimetros = ?,
                Fabricacao = ?,
                Kg_Solicitado = ?,
                Comprado = ?,
                Urgente = ?
            WHERE 
                id = ?
            """,
            (
                bitola, material, norma, dimensao, milimetros, fabricacao, 
                quilos, comprado, urgente, index
            )
        )
        conn.commit()
        conn.close()

    def removeIten(self, index:int) -> None:
        conn = sqlite3.connect(DAVIK_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            f"""
            DELETE FROM DaviK_Main WHERE id={index};
            """
        )
        conn.commit()
        conn.close()

    # out of date
    def makeDictReport(self) -> None:
        with open(davik_report) as file:
            report = json.load(file)
        report = report['report']

        davik = pd.DataFrame(report)

        materiais_agrupados = davik[['bitola','material','norma','dimensao','milimetros','fabricacao','quilos']].groupby(['bitola','material','norma','dimensao','fabricacao']).sum()

        materiais_agrupados.reset_index(inplace=True)
        materiais_dict = materiais_agrupados.to_dict(orient='list')

        materiais_dict = self.marketStandard(materiais_dict)

        materiais_dict['empresa'] = ['DaviK' for x in range(len(materiais_dict['bitola']))]
        with open(davik_report_formated, 'w') as file:
            json.dump(materiais_dict, file, indent=4)

    def makeExcelReport(self) -> None:
        # caminho modelo excel
        modelo_excel = Path(__file__).parent.parent / "resources\\models\\davik_modelo.xlsx"

        # a ser criado
        davik_excel = Path(__file__).parent.parent / "configuration\davik_excel.xlsx"

        conn = sqlite3.connect(DAVIK_DATABASE)
        consulta_sql = ('''
        SELECT 
            Bitola, Material, Norma, Dimensao, Fabricacao, Kg_Solicitado
        FROM DaviK_Main
        WHERE Comprado=0
        ''')

        workbook = load_workbook(modelo_excel)
        sheet = workbook['davik']

        df = pd.read_sql_query(consulta_sql, conn)

        # Adicionar a tabela a partir da célula B8
        start_row = 8
        start_column = 2  # Coluna B

        for row in dataframe_to_rows(df, index=False, header=True):
            for i, value in enumerate(row, start=start_column):
                # if data['urgente'][value]:
                    
                sheet.cell(row=start_row, column=i, value=value)
            start_row += 1

        # Salvar as alterações em arquivo Excel
        workbook.save(davik_excel)
        # self.log_success('Excel konemetal criado com sucesso')

    def receiveItem(self, index, kg_receive, date_receive, fiscal_note):
        # connect to DataBase
        conn = sqlite3.connect(DAVIK_DATABASE)
        cursor = conn.cursor()

        # read database
        cursor.execute(
            f'''
                SELECT Bitola, Material, Norma, Dimensao, Fabricacao,
                Fornecedor, Valor, 
                Kg_Solicitado, Kg_Comprado, 
                Data_Solicitado, Data_Comprado, Data_Previsao, 
                Ordem_Compra

                FROM DaviK_Main WHERE id={index}
            '''
        )
        values = cursor.fetchall()
        
        # set values
        _bit = values[0][0]
        _mat = values[0][1]
        _nor = values[0][2]
        _dim = values[0][3]
        _fab = values[0][4]

        _for = values[0][5]
        _val = values[0][6]

        _kgs = values[0][7]
        _kgc = values[0][8]

        _dts = values[0][9]
        _dtc = values[0][10]
        _dtp = values[0][11]
        
        _oc = values[0][12]

        # write database
        cursor.execute(
            """
            INSERT INTO DaviK_Historic 
            (
                Bitola, Material, Norma, Dimensao, Fabricacao, 
                Fornecedor, Valor,
                Kg_Solicitado, Kg_Comprado, Kg_Recebido, 
                Data_Solicitado, Data_Comprado, Data_Previsao, Data_Recebido, 
                Ordem_Compra, Nota_Fiscal
            ) 

            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                _bit, _mat, _nor, _dim, _fab, 
                _for, _val,
                _kgs, _kgc, kg_receive,
                _dts, _dtc, _dtp, date_receive, 
                _oc, fiscal_note
            )
        )
        conn.commit()


        # delete item database
        cursor.execute(f"DELETE FROM DaviK_Main WHERE id={index}")
        conn.commit()

        # close database
        conn.close()

    def updateHistoric(self, column:str='Data_Recebido', order:str='ASC'):
        conn = sqlite3.connect(DAVIK_DATABASE)
        cursor = conn.cursor()
        cursor.execute(f'''
            SELECT 
                strftime('%d/%m/%Y', Data_Recebido) AS data_formatada,
                Fornecedor, 
                Ordem_Compra, 
                Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                Kg_Recebido, 
                Valor, 
                id
            FROM 
                Davik_Historic
            ORDER BY  {column} {order}''')

        item_list = cursor.fetchall()
        conn.close()
        return item_list

class KonemetalUser(User, LogFileMixin):
    def registerPortfolio(self, Cliente, Ordem_Interna, Pedido, Linha, Produto, Descricao, Quantidade, Entregas, Data_recebido, Data_entrega,  Janela_entrega):



        data_formatada = str(datetime.strptime(data, "%d/%m/%Y").date())
        valor_kg = self.convert_mm_kg(float(milimetros), bitola.upper(), norma.upper(), dimensao.upper())

        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO 
                Konemetal_Portifolio 
                (
                    Cliente,
                    Ordem_Interna,
                    Pedido,
                    Linha,
                    Produto,
                    Descricao,
                    Quantidade,
                    Entregas,
                    Data_recebido,
                    Data_entrega, 
                    Janela_entrega
                ) 
            VALUES 
                (
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                )
            """,
            (
                Cliente,
                Ordem_Interna,
                Pedido,
                Linha,
                Produto,
                Descricao,
                Quantidade,
                Entregas,
                Data_recebido,
                Data_entrega, 
                Janela_entrega
            )
        )
        conn.commit()
        conn.close()

    def registerReport(self, bitola, material, norma, dimensao, milimetros, 
    fabricacao, documento, data, user_name) -> None:
        data_formatada = str(datetime.strptime(data, "%d/%m/%Y").date())
        valor_kg = self.convert_mm_kg(float(milimetros), bitola.upper(), norma.upper(), dimensao.upper())

        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO 
                Konemetal_Main 
                (
                    Bitola, Material, Norma, Dimensao, Milimetros, Fabricacao, 
                    Fornecedor, Valor,
                    Kg_Solicitado, Kg_Comprado,
                    Data_Entrega_Peca, Data_Solicitado, Data_Comprado, Data_Previsao,
                    Ordem_Compra,
                    Urgente, Comprado,
                    Documento, Codigo_Peca, 
                    Solicitante
                ) 
            VALUES 
                (
                    ?, ?, ?, ?, ?, ?,
                    ?, ?, 
                    ?, ?, 
                    ?, ?, ?, ?, 
                    ?, 
                    ?, ?,
                    ?, ?,
                    ?
                )
            """,
            (
                bitola.upper(), material.upper(), norma.upper(), dimensao.upper(), int(milimetros), fabricacao.upper(), 
                None, None,
                valor_kg, None, 
                data_formatada, datetime.now().date(), None, None, 
                None,
                0, 0,
                documento, None,
                user_name
            )
        )
        conn.commit()
        conn.close()

    def saveFile(self):
        shutil.copy(
            KONEMETAL_DATABASE, KONEMETAL_BACKUP
        )

    def editReport(self, id:int, bitola:str, material:str, norma:str, dimensao:str, 
    milimetros:int, fabricacao:str, data_entrega:str, urgente:bool, documento:str, 
    codigo_peca:str) -> None:

        quilos=self.convert_mm_kg(milimetros, bitola, norma, dimensao)

        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            f"""
            UPDATE 
                Konemetal_Main 
            SET 
                Bitola = ?,
                Material = ?,
                Norma = ?,
                Dimensao = ?,
                Milimetros = ?,
                Fabricacao = ?,

                Kg_Solicitado = ?,

                Data_Entrega_Peca = ?,
                Urgente = ?,

                Documento = ?,
                Codigo_Peca = ?
            WHERE 
                id = ?
            """,
            (
                bitola, material, norma, dimensao, milimetros, fabricacao, 
                quilos, data_entrega, urgente, documento, codigo_peca, id
            )
        )
        conn.commit()
        conn.close()

    def removeIten(self, id:int):
        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            f"DELETE FROM Konemetal_Main WHERE id={id}"
        )
        conn.commit()
        conn.close()

    # out of date
    def makeDictReport(self):
        with open(konemetal_report) as file:
            report = json.load(file)
        report = report['report']

        konemetal = pd.DataFrame(report)

        materiais_agrupados = konemetal[['bitola','material','norma','dimensao','milimetros','fabricacao','quilos','data']].groupby(['bitola','material','norma','dimensao','fabricacao']).agg({'milimetros':'sum','quilos':'sum','data':'min'})

        materiais_agrupados.reset_index(inplace=True)
        materiais_dict = materiais_agrupados.to_dict(orient='list')

        materiais_dict = self.marketStandard(materiais_dict)

        materiais_dict['empresa'] = ['Konemetal' for x in range(len(materiais_dict['bitola']))]

        for i in range(len(materiais_dict['bitola'])):
            materiais_dict['data'][i] = self.dataFormater(str(materiais_dict['data'][i]))

        with open(konemetal_report_formated, 'w') as file:
            json.dump(materiais_dict, file, indent=4)


    def makeExcelReport(self):
        # caminho modelo excel
        modelo_excel = Path(__file__).parent.parent / "resources\\models\\konemetal_modelo.xlsx"

        # a ser criado
        konemetal_excel = Path(__file__).parent.parent / "configuration\konemetal_excel.xlsx"

        conn = sqlite3.connect(KONEMETAL_DATABASE)
        consulta_sql = ('''
        SELECT 
            Bitola, Material, Norma, Dimensao, Fabricacao, Kg_Solicitado
        FROM Konemetal_Main
        WHERE comprado=0
        ''')

        workbook = load_workbook(modelo_excel)
        sheet = workbook['konemetal']

        df = pd.read_sql_query(consulta_sql, conn)

        with open(MATERIAL_STANDARDS, 'r', encoding='UTF-8') as file:
            __std = json.load(file) 

        df['Bitola'] = df['Bitola'].replace(__std['Bitola'])
        df['Material'] = df['Material'].replace(__std['Material'])
        df['Norma'] = df['Norma'].replace(__std['Norma'])

        # Adicionar a tabela a partir da célula B8
        start_row = 8
        start_column = 2  # Coluna B

        for row in dataframe_to_rows(df, index=False, header=True):
            for i, value in enumerate(row, start=start_column):
                # if data['urgente'][value]:
                    
                sheet.cell(row=start_row, column=i, value=value)
            start_row += 1

        # Salvar as alterações em arquivo Excel
        workbook.save(konemetal_excel)
        # self.log_success('Excel konemetal criado com sucesso')

    def updateExtraMp(self):
        if self.area == 'Konemetal':
            conn = sqlite3.connect(KONEMETAL_DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
            SELECT Descricao, Quantidade, Documento, Observacao,id
            FROM Konemetal_Extra_Itens
            ''')
            material_list = cursor.fetchall()
            conn.close()
        return material_list

    # out of date
    def registerExtraMp(self, description, amount, document, observation):
        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO 
                Konemetal_Extra_Itens
                (
                    Descricao, Quantidade, Documento, Observacao
                ) 
            VALUES 
                (
                    ?, ?, ?, ?
                )
            """,
            (
                description, amount, document, observation
            )
        )
        conn.commit()
        conn.close()

    # out of date
    def deleteExtraMp(self, index:str) -> None:
        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            f"""
            DELETE FROM Konemetal_Extra_Itens WHERE id={str(index)};
            """
        )
        conn.commit()
        conn.close()

    def receiveItem(self, index, kg_receive, date_receive, fiscal_note):
        # connect to DataBase
        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()

        # read database
        cursor.execute(
            f'''
                SELECT 
                    Bitola, Material, Norma, Dimensao, Fabricacao,
                    Fornecedor, Valor, 
                    Kg_Solicitado, Kg_Comprado, 
                    Data_Entrega_Peca, Data_Solicitado, Data_Comprado, Data_Previsao, 
                    Ordem_Compra, 
                    Documento, Codigo_Peca


                FROM 
                    Konemetal_Main 

                WHERE 
                    id={index}
            '''
        )
        values = cursor.fetchall()
        
        # set values
        _bit = values[0][0]
        _mat = values[0][1]
        _nor = values[0][2]
        _dim = values[0][3]
        _fab = values[0][4]

        _for = values[0][5]
        _val = values[0][6]

        _kgs = values[0][7]
        _kgc = values[0][8]

        _dts = values[0][9]
        _dtc = values[0][10]
        _dtp = values[0][11]
        
        _oc = values[0][12]

        # write database
        cursor.execute(
            """
            INSERT INTO Konemetal_Historic 
            (
                Bitola, Material, Norma, Dimensao, Fabricacao, 
                Fornecedor, Valor,
                Kg_Solicitado, Kg_Comprado, Kg_Recebido, 
                Data_Solicitado, Data_Comprado, Data_Previsao, Data_Recebido, 
                Ordem_Compra, Nota_Fiscal
            ) 

            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                _bit, _mat, _nor, _dim, _fab, 
                _for, _val,
                _kgs, _kgc, kg_receive,
                _dts, _dtc, _dtp, date_receive, 
                _oc, fiscal_note
            )
        )
        conn.commit()


        # delete item database
        cursor.execute(f"DELETE FROM DaviK_Main WHERE id={index}")
        conn.commit()

        # close database
        conn.close()

    def updateHistoric(self, column:str='Data_Recebido', order:str='ASC'):
        # Conectar ao arquivo SQLite
        conn = sqlite3.connect(KONEMETAL_DATABASE)

        # Criar um cursor
        cursor = conn.cursor()

        cursor.execute(f'''
            SELECT 
                Data_Recebido, Fornecedor, Ordem_Compra, Bitola, Material, Norma, Dimensao, 
                Fabricacao, Kg_Recebido, Valor, id
            FROM 
                Konemetal_Historic
            ORDER BY  {column} {order}''')

        # Recuperar os resultados da consulta
        item_list = cursor.fetchall()

        # Iterar pelos item_list
        historic_report = list()
        for item in item_list:
            _new_date = self.dataFormater(item[0])
            _material = f'{item[3]}{item[4]}{item[5]} {item[6]} {item[7][0:3]}'

            historic_report.append(
                [_new_date,
                item[1],
                item[2],
                _material,
                item[8],
                item[9],
                item[10]]
            )

        # Fechar a conexão
        conn.close()
        return historic_report

    def receiveItem(self, index, kg_receive, date_receive, fiscal_note):
        # connect to DataBase
        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()

        # read database
        cursor.execute(
        f'''
            SELECT 
                Bitola, Material, Norma, Dimensao, Fabricacao,
                Fornecedor, Valor, 
                Kg_Solicitado, Kg_Comprado, 
                Data_Entrega_Peca, Data_Solicitado, Data_Comprado, Data_Previsao, 
                Ordem_Compra,
                Documento, Codigo_Peca

            FROM 
                Konemetal_Main 
                
            WHERE
                id={index}
        ''')
        values = cursor.fetchall()
        
        # set values
        _bit = values[0][0]
        _mat = values[0][1]
        _nor = values[0][2]
        _dim = values[0][3]
        _fab = values[0][4]

        _for = values[0][5]
        _val = values[0][6]

        _kgs = values[0][7]
        _kgc = values[0][8]

        _dep = values[0][9]
        _dts = values[0][10]
        _dtc = values[0][11]
        _dtp = values[0][12]
        
        _oc = values[0][13]

        _doc = values[0][14]
        _cod = values[0][15]


        # write database
        cursor.execute(
            """
            INSERT INTO Konemetal_Historic 
            (
                Bitola, Material, Norma, Dimensao, Fabricacao, 
                Fornecedor, Valor,
                Kg_Solicitado, Kg_Comprado, Kg_Recebido, 
                Data_Entrega_Peca, Data_Solicitado, Data_Comprado, Data_Previsao, Data_Recebido, 
                Ordem_Compra, Nota_Fiscal,
                Documento, Codigo_Peca
            ) 

            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                _bit, _mat, _nor, _dim, _fab, 
                _for, _val,
                _kgs, _kgc, kg_receive,
                _dep, _dts, _dtc, _dtp, date_receive, 
                _oc, fiscal_note,
                _doc, _cod
            ))
        conn.commit()


        # delete item database
        cursor.execute(f"DELETE FROM Konemetal_Main WHERE id={index}")
        conn.commit()

        # close database
        conn.close()

class BuyerUser(User, LogFileMixin):

    def makeExcelReport(self):
        # konemetal
        # caminho modelo excel
        modelo_excel = Path(__file__).parent.parent / "resources\\models\\comprador_modelo.xlsx"

        # a ser criado
        retorio_excel = Path(__file__).parent.parent / "configuration\comprador_excel.xlsx"


        conn = sqlite3.connect(KONEMETAL_DATABASE)
        consulta_sql = ('''
        SELECT 
            Bitola, Material, Norma, Dimensao, Fabricacao, Kg_Solicitado
        FROM Konemetal_Main
        WHERE comprado=0
        ''')

        df_kone = pd.read_sql_query(consulta_sql, conn)

        df_kone['Empresa'] = 'Konemetal'



        conn2 = sqlite3.connect(DAVIK_DATABASE)
        consulta_sql2 = ('''
        SELECT 
            Bitola, Material, Norma, Dimensao, Fabricacao, Kg_Solicitado
        FROM DaviK_Main
        WHERE Comprado=0
        ''')

        df_davik = pd.read_sql_query(consulta_sql2, conn2)
        df_davik["Empresa"] = 'Davik'


        df = pd.concat([df_kone, df_davik])


        with open(MATERIAL_STANDARDS, 'r', encoding='UTF-8') as file:
            __std = json.load(file) 

        df['Bitola'] = df['Bitola'].replace(__std['Bitola'])
        df['Material'] = df['Material'].replace(__std['Material'])
        df['Norma'] = df['Norma'].replace(__std['Norma'])

        workbook = load_workbook(modelo_excel)
        sheet = workbook['Comprador']


        # Adicionar a tabela a partir da célula B8
        start_row = 8
        start_column = 2  # Coluna B

        for row in dataframe_to_rows(df, index=False, header=True):
            for i, value in enumerate(row, start=start_column):
                # if data['urgente'][value]:
                    
                sheet.cell(row=start_row, column=i, value=value)
            start_row += 1

        # Salvar as alterações em arquivo Excel
        workbook.save(retorio_excel)
        # self.log_success('Excel konemetal criado com sucesso')

    # out of date
    def makeGeneralReport(self, area:str=''):
        with open(konemetal_report, 'r') as file:
            konemetal = json.load(file)

            ######### POPAR TODOS OS FORJADOS J3 J4 J5 J2-14 T2-14 J6FLANGE
        lista_para_remover = []
        for c in range(len(konemetal['report']['index'])):
            if konemetal['report']['bitola'][c] == 'JL':
                if konemetal['report']['dimensao'][c] in [
                '2CH36','3','4','5','6 FLANGE']:
                    lista_para_remover.append(c)
            if konemetal['report']['bitola'][c] =='JL':
                if konemetal['report']['dimensao'][c]  ==  '2CH36':
                    lista_para_remover.append(c)

        compensador = 0
        for c in lista_para_remover:
            konemetal['report']['index'].pop()
            konemetal['report']['bitola'].pop(c-compensador)
            konemetal['report']['material'].pop(c-compensador)
            konemetal['report']['norma'].pop(c-compensador)
            konemetal['report']['dimensao'].pop(c-compensador)
            konemetal['report']['quilos'].pop(c-compensador)
            konemetal['report']['documento'].pop(c-compensador)
            konemetal['report']['data'].pop(c-compensador)
            compensador += 1
     
        konemetal = konemetal["report"]
        del(konemetal["documento"])
        del(konemetal["data"])

        with open(davik_report, 'r') as file:
            davik = json.load(file)

        davik = davik["report"]

        for i in range(len(davik["index"])):
            konemetal["index"].append(len(konemetal["index"])+1)
            konemetal["bitola"].append(davik["bitola"][i])
            konemetal["material"].append(davik["material"][i])
            konemetal["norma"].append(davik["norma"][i])
            konemetal["dimensao"].append(davik["dimensao"][i])
            konemetal["quilos"].append(davik["quilos"][i])

        relatorio = {
                "index": [],
                "bitola": [],
                "material": [],
                "norma": [],
                "dimensao": [],
                "quilos": []}

        for i in range(len(konemetal["index"])):
            _var = True
            if len(relatorio['index']) == 0:
                relatorio["index"].append(len(relatorio["index"])+1)
                relatorio["bitola"].append(konemetal["bitola"][i])
                relatorio["material"].append(konemetal["material"][i])
                relatorio["norma"].append(konemetal["norma"][i])
                relatorio["dimensao"].append(konemetal["dimensao"][i])
                relatorio["quilos"].append(konemetal["quilos"][i])
            else:
                for c in range(len(relatorio['index'])):
                    if konemetal['bitola'][i] == relatorio['bitola'][c] and konemetal['material'][i] == relatorio['material'][c] and konemetal['norma'][i] == relatorio['norma'][c] and konemetal['dimensao'][i] == relatorio['dimensao'][c]:
                        relatorio['quilos'][c] = int(relatorio['quilos'][c]) + int(konemetal['quilos'][i])
                        _var = False
                if _var:
                    relatorio["index"].append(len(relatorio["index"])+1)
                    relatorio["bitola"].append(konemetal["bitola"][i])
                    relatorio["material"].append(konemetal["material"][i])
                    relatorio["norma"].append(konemetal["norma"][i])
                    relatorio["dimensao"].append(konemetal["dimensao"][i])
                    relatorio["quilos"].append(konemetal["quilos"][i])

        if area == 'view_table':
            relatorio_somado = list()
            for c in range(len(relatorio['index'])):
                material = relatorio["bitola"][c]  +  relatorio["material"][c]  +  relatorio["norma"][c]
                dimensao = relatorio["dimensao"][c]
                quilos = relatorio["quilos"][c]
                relatorio_somado.append([material, dimensao, quilos])
            return relatorio_somado

        elif area == 'view_graph':
            # print(relatorio, type(relatorio))
            return relatorio
            
    # out of date
    def sendEmail(self, relatorio_kone, relatorio_davik, report):
        from automacao import sendEmailAuto
        sendEmailAuto(relatorio_kone, relatorio_davik, report)

    # out of date
    def documentarCompras(self, materiais_comprados):   
        from automacao import documentarComprasAuto
        documentarComprasAuto(materiais_comprados)

    # out of date
    def registerPurchase(self, fornecedor, oc, material_id, quantidade, valor_kg, data_previsao, data, para):
        if para == 'Konemetal':
            conn = sqlite3.connect(KONEMETAL_DATABASE)
            solicitante = 'Konemetal_Main'
        elif para == 'DaviK':

            conn = sqlite3.connect(DAVIK_DATABASE)
            solicitante = 'DaviK_Main'

        if isinstance(valor_kg, str):
            valor_kg = float(valor_kg.replace(',','.'))

        cursor = conn.cursor()
        cursor.execute(f"""
            UPDATE
                {solicitante}
            SET
                Fornecedor = ?, 
                Ordem_Compra = ?, 
                Kg_Comprado = ?, 
                Valor = ?, 
                Data_Comprado = ?,
                Data_Previsao = ?,
                Comprado = ?
            WHERE 
                id = ?
            """,
            (
            fornecedor,
            oc,
            quantidade,
            valor_kg,
            data,
            data_previsao,
            True,
            material_id
            )
        )
        conn.commit()
        conn.close()

    # out of date
    def editBuyedReport(self, numb_index, fornecedor, oc, material, quantidade, valor_kg, para, data):
        with open(materiais_comprados, 'r') as file:
            report_file = json.load(file)
        
        report_file['fornecedor'][numb_index-1] = fornecedor
        report_file['oc'][numb_index-1] = oc
        report_file['material'][numb_index-1] = material
        report_file['quantidade'][numb_index-1] = quantidade
        report_file['valor_kg'][numb_index-1] = valor_kg
        report_file['para'][numb_index-1] = para

        if '/' in data:
            data = data.split('/')
            if len(data) != 3:
                data_formatada = str(datetime.now().year) + str(datetime.now().strftime("%m")) + str(datetime.now().strftime("%d"))
            else:
                if len(data[2]) != 4:
                    data[2] = str(datetime.now().year)
                if len(data[1]) != 2:
                    data[1] = str(datetime.now().strftime("%m"))
                if len(data[0]) != 2:
                    data[0] = str(datetime.now().strftime("%d"))
                data_formatada = data[2] + data[1] + data[0]
        else:
            data_formatada = data

        report_file['data'][numb_index-1] = data_formatada

        with open(materiais_comprados, 'w') as file:
            json.dump(report_file, file, indent=4)

    # out of date
    def removeLineBuyed(self, indice):
        with open(materiais_comprados, 'r') as file:
            report_file = json.load(file)
        
        report_file['indice'].pop()
        report_file['fornecedor'].pop(indice)
        report_file['oc'].pop(indice)
        report_file['material'].pop(indice)
        report_file['quilos'].pop(indice)
        report_file['valor_kg'].pop(indice)
        report_file['para'].pop(indice)
        report_file['data'].pop(indice)

        with open(materiais_comprados, 'w') as file:
            json.dump(report_file, file, indent=4)

    # out of date
    def confirm_receipt(self, index, nota_fiscal):

        with open(historico_compras, 'r') as file:
            historico = json.load(file)
        
        with open(materiais_comprados, 'r') as file:
            comprados = json.load(file)

        dados_combinados = {
            "indice": [len(historico['indice'])+1] + historico['indice'],
            "fornecedor": [comprados["fornecedor"][index]] + historico["fornecedor"],
            "oc": [comprados["oc"][index]] + historico["oc"],
            "material": [comprados["material"][index]] + historico["material"],
            "quilos": [comprados["quilos"][index]] + historico["quilos"],
            "valor_kg": [comprados["valor_kg"][index]] + historico["valor_kg"],
            "para": [comprados["para"][index]] + historico["para"],
            "data": [comprados["data"][index]] + historico["data"],
            'nota_fiscal': [nota_fiscal] + historico["data"]
        }

        with open(historico_compras, 'w') as file:
            json.dump(dados_combinados, file, indent=4)

        self.removeLineBuyed((index))

    # out of date
    def historicFindLine(self, type, info):
        with open(historico_compras, 'r') as file:
            historico = json.load(file)

        relatorio = '{:^13}|{:^6}|{:^20}|{:^11}|{:^9}|{:^11}|{:^9}'.format(
            'FORNECEDOR', 'OC', 'MATERIAL', 'QUANTIDADE', 'VALOR', 'EMPRESA', 'DATA'
            )
        relatorio += f'{"-"*13}|{"-"*6}|{"-"*20}|{"-"*11}|{"-"*9}|{"-"*11}|{"-"*9}'

        if type in ['fornecedor','oc','para']:    
            for i in range(len(historico['indice'])):
                if historico[type][i] == info:
                    #MAKE REPORT
                    relatorio += '{:^13}|{:^6}|{:^20}|{:^11}|{:^9}|{:^11}|{:^9}'.format(
                        historico['fornecedor'][i], historico['oc'][i],
                        historico['material'][i], historico['quantidade'][i],
                        historico['valor_kg'][i], historico['para'][i],
                        historico['data'][i]
                        )

        else:
            for i in range(len(historico['indice'])):
                if historico['material'][i][0] == 'T' \
                or historico['material'][i][0] == 'J' \
                or historico['material'][i][0] == 'C' \
                or historico['material'][i][0] == 'B':
                    
                    bitola = historico['material'][i][:2]
                    norma = historico['material'][i][3]
                    dimensao = historico['material'][i][5:]

                else:
                    bitola = historico['material'][i][0]
                    norma = historico['material'][i][2]
                    dimensao = historico['material'][i][4:]

                if type == 'bitola':
                    if bitola == info:
                        relatorio += '{:^13}|{:^6}|{:^20}|{:^11}|{:^9}|{:^11}|{:^9}'.format(
                            historico['fornecedor'][i], historico['oc'][i],
                            historico['material'][i], historico['quantidade'][i],
                            historico['valor_kg'][i], historico['para'][i],
                            historico['data'][i]
                            )

                elif type == 'norma':
                    if norma == info:
                        relatorio += '{:^13}|{:^6}|{:^20}|{:^11}|{:^9}|{:^11}|{:^9}'.format(
                            historico['fornecedor'][i], historico['oc'][i],
                            historico['material'][i], historico['quantidade'][i],
                            historico['valor_kg'][i], historico['para'][i],
                            historico['data'][i]
                            )

                elif type == 'dimensao':
                    if dimensao == info:
                        relatorio += '{:^13}|{:^6}|{:^20}|{:^11}|{:^9}|{:^11}|{:^9}'.format(
                            historico['fornecedor'][i], historico['oc'][i],
                            historico['material'][i], historico['quantidade'][i],
                            historico['valor_kg'][i], historico['para'][i],
                            historico['data'][i]
                            )

        return relatorio

    def updateHistoric(self, column:str='Data_Recebido', order:str='ASC'):
        if order == 'ASC':
            ascending = True
        else:
            ascending = False

        if column == 'Data_Recebido':
            filter_by = 0
        elif column == 'Fornecedor':
            filter_by = 1
        elif column == 'Ordem_Compra':
            filter_by = 2
        elif column == 'Bitola':
            filter_by = 3
        elif column == 'Kg_Recebido':
            filter_by = 4
        elif column == 'Valor':
            filter_by = 5
        else:
            filter_by = 0  


        conn = sqlite3.connect(DAVIK_DATABASE)
        cursor = conn.cursor()
        cursor.execute(f'''
            SELECT 
                strftime('%d/%m/%Y', Data_Recebido) AS data_formatada,
                Fornecedor, 
                Ordem_Compra, 
                Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                Kg_Recebido, 
                Valor, 
                id
            FROM 
                Davik_Historic
            ORDER BY  {column} {order}''')

        davik_list = cursor.fetchall()
        conn.close()
        
        for i in range(len(davik_list)):
            davik_list[i] = list(davik_list[i])
            davik_list[i].insert(0,'DaviK')


        conn = sqlite3.connect(KONEMETAL_DATABASE)
        cursor = conn.cursor()
        cursor.execute(f'''
            SELECT 
                strftime('%d/%m/%Y', Data_Recebido) AS data_formatada,
                Fornecedor, 
                Ordem_Compra, 
                Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                Kg_Recebido, 
                Valor, 
                id
            FROM 
                Konemetal_Historic
            ORDER BY  {column} {order}''')

        konemetal_list = cursor.fetchall()
        conn.close()

        for i in range(len(konemetal_list)):
            konemetal_list[i] = list(konemetal_list[i])
            konemetal_list[i].insert(0,'Konemetal')

        item_list = konemetal_list + davik_list

        df = pd.DataFrame(item_list)

        df[1] = pd.to_datetime(df[1], format='%d/%m/%Y')

        order_list_mp = df.sort_values(by=filter_by, ascending=ascending)

        order_list_mp[1] = order_list_mp[1].dt.strftime('%d/%m/%Y')

        order_list_mp = order_list_mp.values.tolist()

        return order_list_mp

    def filterReport(self, filter_by:str, filter_key:str, historic:None=None):

        if historic:
            conn = sqlite3.connect(DAVIK_DATABASE)
            cursor = conn.cursor()
            cursor.execute(f'''
                SELECT 
                    strftime('%d/%m/%Y', Data_Recebido) AS data_formatada,
                    Fornecedor, 
                    Ordem_Compra, 
                    Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                    Kg_Recebido, 
                    Valor, 
                    id
                FROM 
                    Davik_Historic
                WHERE 
                    "{filter_by}" = "{filter_key}"
                ORDER BY  
                    data_formatada ASC''')

            davik_list = cursor.fetchall()
            conn.close()
            # print('\n\n\ndavik',davik_list)
            
            for i in range(len(davik_list)):
                davik_list[i] = list(davik_list[i])
                davik_list[i].insert(0,'DaviK')


            conn = sqlite3.connect(KONEMETAL_DATABASE)
            cursor = conn.cursor()
            cursor.execute(f'''
                SELECT 
                    strftime('%d/%m/%Y', Data_Recebido) AS data_formatada,
                    Fornecedor, 
                    Ordem_Compra, 
                    Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                    Kg_Recebido, 
                    Valor, 
                    id
                FROM 
                    Konemetal_Historic
                WHERE 
                    "{filter_by}" = "{filter_key}"
                ORDER BY  
                    data_formatada ASC''')

            konemetal_list = cursor.fetchall()
            conn.close()
            # print('\n\n\nKonemetal',konemetal_list)

            for i in range(len(konemetal_list)):
                konemetal_list[i] = list(konemetal_list[i])
                konemetal_list[i].insert(0,'Konemetal')

            item_list = konemetal_list + davik_list

            # print('\n\n\ntotal',item_list)

            df = pd.DataFrame(item_list)
            
            # df.head()
            
            df[1] = pd.to_datetime(df[1], format='%d/%m/%Y')

            order_list_mp = df.sort_values(by=[1])

            order_list_mp[1] = order_list_mp[1].dt.strftime('%d/%m/%Y')

            order_list_mp = order_list_mp.values.tolist()
        else:
            conn = sqlite3.connect(DAVIK_DATABASE)
            cursor = conn.cursor()
            cursor.execute(f'''
                SELECT 
                    Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                    Kg_Solicitado, 
                    strftime('%d/%m/%Y', Data_Solicitado) AS data_formatada,
                    Valor, 
                    Data_Previsao, 
                    id
                FROM 
                    Davik_Main
                WHERE 
                    "{filter_by}" = "{filter_key}"
                ORDER BY  
                    data_formatada ASC''')

            davik_list = cursor.fetchall()
            conn.close()
            # print('\n\n\ndavik',davik_list)
            
            for i in range(len(davik_list)):
                davik_list[i] = list(davik_list[i])
                davik_list[i].insert(0,'DaviK')


            conn = sqlite3.connect(KONEMETAL_DATABASE)
            cursor = conn.cursor()
            cursor.execute(f'''
                SELECT 
                    Bitola || '' || Material || '' || Norma || ' ' || Dimensao || ' ' || Fabricacao AS Material_format,
                    Kg_Solicitado, 
                    strftime('%d/%m/%Y', Data_Solicitado) AS data_formatada,
                    Valor, 
                    Data_Previsao, 
                    id
                FROM 
                    Konemetal_Main
                WHERE 
                    "{filter_by}" = "{filter_key}"
                ORDER BY  
                    data_formatada ASC''')

            konemetal_list = cursor.fetchall()
            conn.close()
            # print('\n\n\nKonemetal',konemetal_list)

            for i in range(len(konemetal_list)):
                konemetal_list[i] = list(konemetal_list[i])
                konemetal_list[i].insert(0,'Konemetal')

            item_list = konemetal_list + davik_list

            # print('\n\n\ntotal',item_list)

            df = pd.DataFrame(item_list)
            
            # df.head()
            
            df[3] = pd.to_datetime(df[3], format='%d/%m/%Y')

            order_list_mp = df.sort_values(by=[3])

            order_list_mp[3] = order_list_mp[3].dt.strftime('%d/%m/%Y')

            order_list_mp = order_list_mp.values.tolist()

        return order_list_mp

    def unBuy(self, id:int, area:str) -> None:
        if area == 'Konemetal':
            conn = sqlite3.connect(KONEMETAL_DATABASE)
            solicitante = 'Konemetal_Main'

        elif area == 'DaviK':
            conn = sqlite3.connect(DAVIK_DATABASE)
            solicitante = 'DaviK_Main'

        cursor = conn.cursor()
        cursor.execute(f"""
            UPDATE
                {solicitante}
            SET
                Fornecedor = NULL,
                Ordem_Compra = NULL,
                Kg_Comprado = NULL,
                Valor = NULL,
                Data_Comprado = NULL,
                Data_Previsao = NULL,
                Comprado = ?
            WHERE 
                id = ?
            """,
            (
                False,
                id
            )
        )
        conn.commit()
        conn.close()

if __name__ == '__main__':
    print(datetime.now().date().strftime('%d/%m/%Y'))
    print(datetime.now().date())
    data = datetime.strptime('2023-11-17', '%Y-%m-%d').strftime('%d/%m/%Y')
    dia_hoje = datetime.now().date().strftime('%d/%m/%Y')
    print(data > dia_hoje)

    # print(dia_hoje)


    # li1 = [[1,2,3],[1,2,3]]
    # li2 = [[4,5,6],[4,5,6]]
    # li3 = li1 + li2

    # print(li3)

    # davik = DavikUser('Gessica', '12345', 'DaviK', 'pc', 
    # True, True, True, True, True, 
    # True, True, True, True, 
    # True, True, 'theme_name:str', "button_color:str", 
    # 'pressed_button_color:str', 'second_color:str', 'text_color:str', 'font:str', 
    # 10, [], [], [],
    # {})
    # print(davik.filterReport('Bitola','R'))
    # ...
